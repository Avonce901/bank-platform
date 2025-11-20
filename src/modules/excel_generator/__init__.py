"""
Excel Generator Module
Creates Excel reports from data
"""
import io
import json
from typing import Dict, List, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Generate Excel files from data"""

    def __init__(self):
        self.workbook = None
        self.worksheet = None

    def create_report(self, title: str, data: List[Dict[str, Any]], 
                     filename: str = None) -> io.BytesIO:
        """
        Create an Excel report from data
        
        Args:
            title: Report title
            data: List of dictionaries with data
            filename: Optional filename for the report
            
        Returns:
            BytesIO object containing Excel file
        """
        try:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Report"
            
            # Add title
            self.worksheet['A1'] = title
            title_cell = self.worksheet['A1']
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            self.worksheet.merge_cells('A1:Z1')
            self.worksheet.row_dimensions[1].height = 25
            
            # Add timestamp
            self.worksheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            self.worksheet['A2'].font = Font(italic=True, size=10, color="666666")
            
            if not data:
                return self._save_workbook()
            
            # Get headers from first row
            headers = list(data[0].keys())
            
            # Write headers
            header_row = 4
            for col_idx, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=header_row, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data rows
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row_idx, row_data in enumerate(data, header_row + 1):
                for col_idx, header in enumerate(headers, 1):
                    cell = self.worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = row_data.get(header, "")
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    # Alternate row colors
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                self.worksheet.column_dimensions[column_letter].width = max(15, len(str(header)) + 5)
            
            return self._save_workbook()
        except Exception as e:
            raise ValueError(f"Failed to create Excel report: {str(e)}")

    def create_financial_report(self, accounts: List[Dict[str, Any]], 
                               transactions: List[Dict[str, Any]]) -> io.BytesIO:
        """
        Create a financial report with accounts and transactions
        
        Args:
            accounts: List of account dictionaries
            transactions: List of transaction dictionaries
            
        Returns:
            BytesIO object containing Excel file
        """
        try:
            self.workbook = openpyxl.Workbook()
            
            # Create Accounts sheet
            ws_accounts = self.workbook.active
            ws_accounts.title = "Accounts"
            self._write_sheet("Accounts Summary", accounts, ws_accounts)
            
            # Create Transactions sheet
            ws_transactions = self.workbook.create_sheet("Transactions")
            self._write_sheet("Transaction History", transactions, ws_transactions)
            
            # Create Summary sheet
            ws_summary = self.workbook.create_sheet("Summary", 0)
            self._write_summary(ws_summary, accounts, transactions)
            
            return self._save_workbook()
        except Exception as e:
            raise ValueError(f"Failed to create financial report: {str(e)}")

    def _write_sheet(self, title: str, data: List[Dict[str, Any]], worksheet) -> None:
        """Write data to worksheet"""
        worksheet['A1'] = title
        title_cell = worksheet['A1']
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        worksheet.merge_cells('A1:Z1')
        
        if not data:
            return
        
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        
        for row_idx, row_data in enumerate(data, 4):
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(header, "")

    def _write_summary(self, worksheet, accounts: List[Dict[str, Any]], 
                      transactions: List[Dict[str, Any]]) -> None:
        """Write summary statistics"""
        worksheet['A1'] = "Financial Summary"
        worksheet['A1'].font = Font(size=14, bold=True)
        
        # Account summary
        worksheet['A3'] = "Total Accounts:"
        worksheet['B3'] = len(accounts)
        
        total_balance = sum(float(acc.get('balance', 0)) for acc in accounts if isinstance(acc.get('balance'), (int, float, str)))
        worksheet['A4'] = "Total Balance:"
        worksheet['B4'] = f"${total_balance:,.2f}"
        
        # Transaction summary
        worksheet['A6'] = "Total Transactions:"
        worksheet['B6'] = len(transactions)

    def _save_workbook(self) -> io.BytesIO:
        """Save workbook to BytesIO"""
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output


def generate_excel(data: List[Dict[str, Any]], title: str = "Report") -> io.BytesIO:
    """
    Main function to generate Excel file
    
    Args:
        data: List of dictionaries with data
        title: Report title
        
    Returns:
        BytesIO object containing Excel file
    """
    generator = ExcelGenerator()
    return generator.create_report(title, data)


def generate_financial_report(accounts: List[Dict[str, Any]], 
                             transactions: List[Dict[str, Any]]) -> io.BytesIO:
    """
    Generate financial report
    
    Args:
        accounts: List of account data
        transactions: List of transaction data
        
    Returns:
        BytesIO object containing Excel file
    """
    generator = ExcelGenerator()
    return generator.create_financial_report(accounts, transactions)
